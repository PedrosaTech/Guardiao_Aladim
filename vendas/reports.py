"""
Rotinas de relatório de vendas (consolidado).

- Base: ItemPedidoVenda + PedidoVenda (somente FATURADO)
- Export: Excel (openpyxl) e PDF (WeasyPrint)
"""
from __future__ import annotations

from dataclasses import dataclass
from decimal import Decimal
from typing import Any, Dict, Iterable, List, Optional, Tuple

from django.db.models import Count, F, Sum
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone

from produtos.models import CodigoBarrasAlternativo
from vendas.models import ItemPedidoVenda

try:
    from weasyprint import HTML

    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False


@dataclass(frozen=True)
class TotaisRelatorio:
    total_quantidade: Decimal
    total_valor: Decimal
    total_desconto: Decimal
    total_pedidos: int
    total_produtos: int


def queryset_base_vendas():
    return (
        ItemPedidoVenda.objects.filter(
            is_active=True,
            pedido__is_active=True,
            pedido__status="FATURADO",
        )
        .select_related(
            "pedido",
            "pedido__loja",
            "pedido__cliente",
            "produto",
            "produto__categoria",
            "codigo_alternativo_usado",
            "codigo_alternativo_usado__fornecedor",
        )
    )


def aplicar_filtros(qs, form):
    if not form.is_valid():
        return qs.none()

    cd = form.cleaned_data
    if cd.get("data_inicio"):
        qs = qs.filter(pedido__data_emissao__date__gte=cd["data_inicio"])
    if cd.get("data_fim"):
        qs = qs.filter(pedido__data_emissao__date__lte=cd["data_fim"])

    if cd.get("produto"):
        qs = qs.filter(produto=cd["produto"])
    if cd.get("categoria"):
        qs = qs.filter(produto__categoria=cd["categoria"])
    if cd.get("loja"):
        qs = qs.filter(pedido__loja=cd["loja"])
    if cd.get("cliente"):
        qs = qs.filter(pedido__cliente=cd["cliente"])
    if cd.get("classe_risco"):
        qs = qs.filter(produto__classe_risco=cd["classe_risco"])

    fornecedor = cd.get("fornecedor")
    if fornecedor:
        qs = qs.filter(
            codigo_alternativo_usado__isnull=False,
            codigo_alternativo_usado__fornecedor=fornecedor,
        )

    return qs


def calcular_totais(qs) -> TotaisRelatorio:
    agg = qs.aggregate(
        total_quantidade=Sum("quantidade"),
        total_valor=Sum("total"),
        total_desconto=Sum("desconto"),
        total_pedidos=Count("pedido_id", distinct=True),
        total_produtos=Count("produto_id", distinct=True),
    )
    return TotaisRelatorio(
        total_quantidade=agg["total_quantidade"] or Decimal("0.000"),
        total_valor=agg["total_valor"] or Decimal("0.00"),
        total_desconto=agg["total_desconto"] or Decimal("0.00"),
        total_pedidos=int(agg["total_pedidos"] or 0),
        total_produtos=int(agg["total_produtos"] or 0),
    )


def top_produtos(qs, limit: int = 10):
    return list(
        qs.values("produto_id", "produto__codigo_interno", "produto__descricao")
        .annotate(
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
        )
        .order_by("-quantidade")[:limit]
    )


def codigos_alternativos_info(produto_id: Optional[int]) -> List[Dict[str, Any]]:
    if not produto_id:
        return []
    out = []
    qs = (
        CodigoBarrasAlternativo.objects.filter(produto_id=produto_id, is_active=True)
        .select_related("fornecedor")
        .order_by("codigo_barras")
    )
    for c in qs:
        out.append(
            {
                "codigo": c.codigo_barras,
                "descricao": c.descricao,
                "fornecedor": c.fornecedor.razao_social if c.fornecedor else None,
                "multiplicador": c.multiplicador,
            }
        )
    return out


def agregar(qs, agrupar_por: str, ordenar_por: str):
    # Para simplificar ordenação, normalizamos campos “nome/valor_total/quantidade”
    if agrupar_por == "produto":
        base = qs.values(
            "produto_id",
            "produto__codigo_interno",
            "produto__descricao",
            "produto__categoria__nome",
            "produto__classe_risco",
        ).annotate(
            nome=F("produto__descricao"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by(ordenar_por)

    if agrupar_por == "categoria":
        base = qs.values(
            "produto__categoria_id",
            "produto__categoria__nome",
        ).annotate(
            nome=F("produto__categoria__nome"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            produtos_count=Count("produto_id", distinct=True),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by(ordenar_por)

    if agrupar_por == "fornecedor":
        base = qs.values(
            "codigo_alternativo_usado__fornecedor_id",
            "codigo_alternativo_usado__fornecedor__razao_social",
        ).annotate(
            nome=F("codigo_alternativo_usado__fornecedor__razao_social"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            produtos_count=Count("produto_id", distinct=True),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by(ordenar_por)

    if agrupar_por == "cliente":
        base = qs.values(
            "pedido__cliente_id",
            "pedido__cliente__nome_razao_social",
        ).annotate(
            nome=F("pedido__cliente__nome_razao_social"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            produtos_count=Count("produto_id", distinct=True),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by(ordenar_por)

    if agrupar_por == "dia":
        base = qs.annotate(data=TruncDate("pedido__data_emissao")).values("data").annotate(
            nome=F("data"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            produtos_count=Count("produto_id", distinct=True),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by("-data")

    if agrupar_por == "mes":
        base = qs.annotate(data=TruncMonth("pedido__data_emissao")).values("data").annotate(
            nome=F("data"),
            quantidade=Sum("quantidade"),
            valor_total=Sum("total"),
            produtos_count=Count("produto_id", distinct=True),
            pedidos_count=Count("pedido_id", distinct=True),
        )
        return base.order_by("-data")

    return []


def exportar_excel(dados: List[Dict[str, Any]], totais: TotaisRelatorio, filtros_desc: str) -> HttpResponse:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse(
            "Biblioteca openpyxl não instalada. Execute: pip install openpyxl",
            status=500,
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Vendas"

    ws.merge_cells("A1:F1")
    ws["A1"] = "RELATÓRIO DE VENDAS (FATURADO)"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A3"] = "Filtros:"
    ws["B3"] = filtros_desc

    ws["A5"] = "Total Quantidade:"
    ws["B5"] = float(totais.total_quantidade)
    ws["A6"] = "Total Valor:"
    ws["B6"] = float(totais.total_valor)
    ws["A7"] = "Total Pedidos:"
    ws["B7"] = totais.total_pedidos

    row = 9
    headers = ["Nome", "Quantidade", "Valor Total", "Pedidos", "Produtos"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    row += 1

    for item in dados:
        ws.cell(row=row, column=1, value=str(item.get("nome") or "—"))
        ws.cell(row=row, column=2, value=float(item.get("quantidade") or 0))
        ws.cell(row=row, column=3, value=float(item.get("valor_total") or 0))
        ws.cell(row=row, column=4, value=int(item.get("pedidos_count") or 0))
        ws.cell(row=row, column=5, value=int(item.get("produtos_count") or 0))
        row += 1

    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 24 if col == 1 else 16

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="relatorio_vendas.xlsx"'
    wb.save(resp)
    return resp


def exportar_pdf(request, dados: List[Dict[str, Any]], totais: TotaisRelatorio, context_extra: Dict[str, Any]) -> HttpResponse:
    if not WEASYPRINT_AVAILABLE:
        return HttpResponse(
            "<h1>Erro: WeasyPrint não instalado</h1><pre>pip install weasyprint</pre>",
            status=500,
        )

    context = {
        **context_extra,
        "dados": dados,
        "totais": totais,
        "data_geracao": timezone.now(),
    }
    html_string = render_to_string("vendas/relatorio_vendas_pdf.html", context)
    base_url = request.build_absolute_uri("/")
    pdf = HTML(string=html_string, base_url=base_url).write_pdf()

    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = 'attachment; filename="relatorio_vendas.pdf"'
    return resp

